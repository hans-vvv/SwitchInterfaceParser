""" SwitchInterfaceParser script
This script reads interface and vlan configuration parts of Cisco IOS
multilayer switches and prints all items in an excel sheet.
Using excel features like autofilter you can analyse the (vlan)interface
and vlan specific configuration items.
Multiple switch configurations can be read.
Set the directory where the configurations reside at your needs in the
main function of the script.
"""

import re
from collections import defaultdict
import glob
from openpyxl.utils import get_column_letter
from openpyxl import Workbook



class ReSearcher:

    """
    Helper class to enable evaluation
    and regex formatting in a single line
    """

    match = None

    def __call__(self, pattern, string):
        self.match = re.search(pattern, string)
        return self.match

    def __getattr__(self, name):
        return getattr(self.match, name)


def xlref(row, column, zero_indexed=True):

    """
    openpyxl helper
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def splitrange(raw_range):

    """
    ex. splitrange('105-107') will return ['105','106','107']
    """

    result = []

    if  re.search(r'^(\d+)\-(\d+)$', raw_range):
        match = re.search(r'^(\d+)\-(\d+)$', raw_range)
        first = int(format(match.group(1)))
        last = int(format(match.group(2)))
        for i in range(first, last+1):
            result.append(str(i))
        return result


def get_value(key, item):

    """
    key + value = item
    function return value for given key and item
    """

    if key.strip() == item.strip():
        return key
    else:
        item = item.lstrip()
        result = re.search('^('+key+')(.*)', item)
        value = format(result.group(2)).lstrip()
        return value


def get_key(item, key_length):

    """
    key + value = item
    number of words of key = key_length
    function returns key
    """

    word = item.strip().split()
    if key_length == 0: # fix
        return item
    elif len(word) == key_length:
        return item
    else:
        return ' '.join(word[0:key_length])


def get_switch_info(configfiles):

    """
    This function stores interface and Vlan specific parts of switch
    configurations in a (nested) dictionary.
    For the following keys, lists are used to store values:
    - switchport trunk allowed vlan (add)
    - standby (HSRP)
    - ip helper-address
    Known Caveats:
    - No support for secundairy ip addresses
    - Only Vlan name is read under VLAN configuration
    - If SVI of VLAN exists but VLAN doesn't then VLAN index appears to be
      present in result.
    """

    # Helper function
    def store_intf_items(line, portinfo, vlan, intf):

        """
        This helper function stores interface items.
        The following methods are used in the given order to determine
        which part of an interface item is considered to be a key and which
        part a value.

        1. First portkey_exceptions list is examined. If an interface item
           contains the words found in this list then key = item in the list
           and value = remaining words of the interface item. If an interface
           item is found then the other methods are not considered.
        2. Portkeys dict is examined. If interface item contains an item
           found in a list of the dict then corresponding key (i.e. 1 or 2)
           is used to split the item. The key of the item is equal to the
           number of words of the dict key, the rest of the item = value.
           Example: If line = channel-group 2 mode active, then
           key = "channel-group"  and value = "2 mode active". If an interface
           item is found then the last method is not considered.
        3. Default method. Last word of line = value
           and all preceding words = key.
        """

        portkey_exceptions = ['ip vrf forwarding']

        portkeys = {1: ['hold-queue', 'standby', 'channel-group', 'description'],
                    2: ['switchport port-security', 'ip', 'spanning-tree',
                        'speed auto', 'srr-queue bandwidth']}

        line = line.lstrip()
        found_item = False

        # 'Method 1'
        for key in portkey_exceptions:
            if key in line:
                if 'Vlan' in intf:
                    vlaninfo[vlan][key] = get_value(key, line)
                    found_item = True
                else:
                    portinfo[intf][key] = get_value(key, line)
                    found_item = True

        # 'Method 2'
        for key_length in portkeys:
            if found_item:
                continue
            for item in portkeys[key_length]:
                if item not in line:
                    continue
                key = get_key(line, key_length)
                if 'standby' in line:
                    if 'Vlan' in intf:
                        standby.append(get_value(key, line))
                        vlaninfo[vlan]['standby'] = ','.join(standby)
                        found_item = True
                    else:
                        standby.append(get_value(key, line))
                        portinfo[intf]['standby'] = ','.join(standby)
                        found_item = True
                elif 'ip helper-address' in line:
                    if 'Vlan' in intf:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        vlaninfo[vlan]['ip helper-address'] = helper
                        found_item = True
                    else:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        portinfo[intf]['ip helper-address'] = helper
                        found_item = True
                elif 'Vlan' in intf:
                    vlaninfo[vlan][key] = get_value(key, line)
                    found_item = True
                else:
                    portinfo[intf][key] = get_value(key, line)
                    found_item = True

        # 'Method 3 or default method'
        if not found_item:
            key = get_key(line, len(word)-1)
            if 'switchport trunk allowed vlan' in line:
                for raw_vlans in get_value(key, line).split(','):
                    if '-' in raw_vlans:
                        for vlan_id in splitrange(raw_vlans):
                            vlan_allow_list.append(vlan_id)
                    else:
                        vlan_allow_list.append(raw_vlans)
                portinfo[intf]['vlan_allow_list'] = ','.join(vlan_allow_list)
            elif 'Vlan' in intf:
                vlaninfo[vlan][key] = get_value(key, line)
            else:
                portinfo[intf][key] = get_value(key, line)


        return portinfo


    # Start main part of function
    switchinfo = defaultdict(dict) # Dict containing all info

    for configfile in configfiles:

        with open(configfile, 'r') as lines:

            portinfo = defaultdict(dict)
            vlaninfo = defaultdict(dict)
            scanfile = False

            match = ReSearcher()

            for line in lines:

                line = line.rstrip()
                word = line.split()


                if match(r'^interface Vlan(\d+)', line):
                    intf = format(match.group(0))
                    vlan = format(match.group(1))
                    vlaninfo[vlan]['vlan_id'] = vlan
                    scanfile = True
                    standby = []
                    ip_helper = []

                elif match(r'^vlan (\d+)\-(\d+)$', line):
                    start_vlan = int(match.group(1))
                    stop_vlan = int(match.group(2))
                    for vlan in range(start_vlan, stop_vlan+1):
                        vlaninfo[str(vlan)]['vlan_id'] = str(vlan)

                elif match(r'^vlan (\d+)$', line):
                    vlan = format(match.group(1))
                    vlaninfo[vlan]['vlan_id'] = vlan

                elif match(r'^interface (.*)', line):
                    intf = format(match.group(1))
                    scanfile = True
                    vlan_allow_list = []
                    standby = []
                    ip_helper = []

                elif match(r'^ name (.*)', line):
                    vlaninfo[vlan]['name'] = format(match.group(1))

                elif match(r'^ no (.*)', line) and scanfile:
                    key = format(match.group(1))
                    value = format(match.group(0))
                    if 'Vlan' in intf:
                        vlaninfo[vlan][key] = value
                    else:
                        portinfo[intf][key] = value

                elif match(r'^hostname (.*)', line):
                    hostname = format(match.group(1))

                elif match(r'^ip forward-protocol nd', line):
                    scanfile = False

                elif match(r'^(ip classless|ip default-gateway)', line):
                    scanfile = False

                # interface items are stored with helper function
                elif match('^ .*', line) and scanfile:
                    store_intf_items(line, portinfo, vlan, intf)


            switchinfo[hostname]['portinfo'] = portinfo
            switchinfo[hostname]['vlaninfo'] = vlaninfo

    return switchinfo


def info_to_xls(switchinfo):

    """
    Function print Switchinfo dictionaty to excel file. Vlan
    are port info are printed in seperated tabs.
    """

    # Calculate list of keys to be present in Excel sheets
    vlankeys = []
    portkeys = []

    for hostname in switchinfo:
        for vlanid in switchinfo[hostname]['vlaninfo']:
            for key in switchinfo[hostname]['vlaninfo'][vlanid]:
                vlankeys.append(key)

    for hostname in switchinfo:
        for intf in switchinfo[hostname]['portinfo']:
            for key in switchinfo[hostname]['portinfo'][intf]:
                portkeys.append(key)

    vlankeys = sorted(set(vlankeys))
    portkeys = sorted(set(portkeys))
    vlankeys.remove('vlan_id')
    vlankeys.remove('name')
    portkeys.remove('description')
    vlankeys.insert(0, 'name')
    portkeys.insert(0, 'description')

    wb = Workbook()
    ws = wb.create_sheet("Vlaninfo", 0)
    ws = wb.create_sheet("Portinfo", 0)

    ws = wb['Vlaninfo']

    count_vlan_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'vlan_id'
    for count, vlanitem in enumerate(vlankeys):
        ws[xlref(0, count+2)] = vlanitem

    for hostname in switchinfo:
        for vlan, vlanitems in switchinfo[hostname]['vlaninfo'].items():
            ws[xlref(count_vlan_row+1, 0)] = hostname
            ws[xlref(count_vlan_row+1, 1)] = vlan

            for count_col, vlanitem in enumerate(vlankeys):
                vlankey = vlanitems.get(vlankeys[count_col], '')
                ws[xlref(count_vlan_row+1, count_col+2)] = vlankey
            count_vlan_row += 1

    ws = wb['Portinfo']

    count_port_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'interface'
    for count, portitem in enumerate(portkeys):
        ws[xlref(0, count+2)] = portitem

    for hostname in switchinfo.keys():
        for intf, intf_items in switchinfo[hostname]['portinfo'].items():
            ws[xlref(count_port_row+1, 0)] = hostname
            ws[xlref(count_port_row+1, 1)] = intf

            for count_col, portitem in enumerate(portkeys):
                portkey = intf_items.get(portkeys[count_col], '')
                ws[xlref(count_port_row+1, count_col+2)] = portkey
            count_port_row += 1

    wb.save('result.xlsx')


if __name__ == "__main__":

    #os.chdir('C:/Users/Hans/Desktop/GIT/SwitchInterfaceParser')

    configfiles = [file for file in glob.glob('*.txt')]

    # Retrieve interface and vlan info from configuration file
    # and store in switchinfo object.
    switchinfo = get_switch_info(configfiles)

    # Print Switchinfo object in excel file.
    info_to_xls(switchinfo)
