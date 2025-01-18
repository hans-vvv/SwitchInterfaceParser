from openpyxl import Workbook
from utils import xlref


def xls_writer(tree):
    """
    Function print Switchinfo dictionary to excel file. Vlan. port info and
    static routes are printed in seperated tabs.
    """
    # Calculate set of keys.
    vlankeys = set()
    portkeys = set()

    for switch in tree:
        for vlan, vlanitems in switch['vlan'].items():
            vlankeys.update(vlanitems.keys())
        for port, portitems in switch['port'].items():
            portkeys.update(portitems.keys())
                
    vlankeys = sorted(set(vlankeys))
    portkeys = sorted(set(portkeys))
    if 'vlanindex' in vlankeys:
        vlankeys.remove('vlanindex')
    if 'name' in vlankeys:
        vlankeys.remove('name')
    if 'description' in portkeys:
        portkeys.remove('description')
    if 'portindex' in portkeys:
        portkeys.remove('portindex')
    vlankeys.insert(0, 'name')
    portkeys.insert(0, 'description')

    wb = Workbook()
    wb.create_sheet("Vlaninfo", 0)
    wb.create_sheet("Portinfo", 0)
    wb.create_sheet('StaticRoutes')

    ws = wb['Vlaninfo']

    count_vlan_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'vlanindex'
    for count, vlankey in enumerate(vlankeys):
        ws[xlref(0, count+2)] = vlankey

    for switch in tree:
        for vlan, vlanitems in switch['vlan'].items():
            ws[xlref(count_vlan_row+1, 0)] = switch['hostname']
            ws[xlref(count_vlan_row+1, 1)] = vlan

            for count_col, vlankey in enumerate(vlankeys):
                value = vlanitems.get(vlankey, '')
                if isinstance(value, list):
                    value = ','.join(value)
                ws[xlref(count_vlan_row+1, count_col+2)] = value
            count_vlan_row += 1

    ws = wb['Portinfo']

    count_port_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'interface'
    for count, portkey in enumerate(portkeys):
        ws[xlref(0, count+2)] = portkey

    for switch in tree:
        for port, portitems in switch['port'].items():
            ws[xlref(count_port_row+1, 0)] = switch['hostname']
            ws[xlref(count_port_row+1, 1)] = port

            for count_col, portkey in enumerate(portkeys):
                value = portitems.get(portkey, '')
                if isinstance(value, list):
                    value = ','.join(value)
                ws[xlref(count_port_row+1, count_col+2)] = value
            count_port_row += 1

    ws = wb['StaticRoutes']
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'subnet'
    ws[xlref(0, 2)] = 'mask'
    ws[xlref(0, 3)] = 'next-hop'
    ws[xlref(0, 4)] = 'vrf'

    index = 0
    for switch in tree:
        for route in switch['static route']:
            ws[xlref(index + 1, 0)] = switch['hostname']
            ws[xlref(index + 1, 1)] = route[1]
            ws[xlref(index + 1, 2)] = route[2]
            ws[xlref(index + 1, 3)] = route[3]
            ws[xlref(index + 1, 4)] = route[0]
            index += 1

    wb.save('result.xlsx')
