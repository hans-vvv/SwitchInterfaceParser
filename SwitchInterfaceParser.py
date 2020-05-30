""" SwitchInterfaceParser script
This script reads interface and vlan configuration parts of Cisco IOS
multilayer switches and prints all items in an excel sheet.
Using excel features like autofilter you can analyse the (vlan)interface
and vlan specific configuration items.
Multiple switch configurations can be read.
Set the directory where the configurations reside at your needs in the
main function of the script.
"""

import re
import json
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ReSearcher:

    """
    Helper class to enable evaluation
    and regex formatting in a single line
    """
    match = None

    def __call__(self, pattern, string):
        self.match = re.search(pattern, string)
        return self.match

    def __getattr__(self, name):
        return getattr(self.match, name)


class Vividict(dict):
    """" Helper to constuct nested dicts """
    def __missing__(self, key):
        value = self[key] = type(self)()
        return value

    def __str__(self):
        return json.dumps(self, indent=4)


def xlref(row, column, zero_indexed=True):

    """
    openpyxl helper
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def splitrange(raw_range):

    """
    ex. splitrange('105-107') will return ['105','106','107']
    """

    m = re.search(r'^(\d+)\-(\d+)$', raw_range)
    if m:
        first = int(format(m.group(1)))
        last = int(format(m.group(2)))
        return [str(i) for i in range(first, last+1)]


def get_value(key, item):

    """
    key + value = item
    function returns value for given key and item
    """

    if key.strip() == item.strip():
        return key
    else:
        item = item.lstrip()
        result = re.search('^('+key+')(.*)', item)
        value = format(result.group(2)).lstrip()
        return value


def get_key(item, key_length):

    """
    key + value = item
    number of words of key = key_length
    function returns key
    """

    word = item.strip().split()
    if key_length == 0: # fix
        return item
    elif len(word) == key_length:
        return item
    else:
        return ' '.join(word[0:key_length])


def get_switch_info(configfiles):

    """
    This function stores interface and Vlan specific parts of switch
    configurations in a list of nested dictionaries.
    For the following keys, lists are used to store values:
    - switchport trunk allowed vlan (add)
    - standby (HSRP)
    - ip helper-address
    Known Caveats:
    - No support for secundairy ip addresses
    - Only Vlan name is read under VLAN configuration
    - If SVI of VLAN exists but VLAN doesn't then VLAN index appears to be
      present in result.
    """

    # Helper function
    def store_port_items(line, vlanindex, portindex):

        """
        This helper function stores interface items.
        The following methods are used in the given order to determine
        which part of an interface item is considered to be a key and which
        part a value.

        1. First portkey_exceptions list is examined. If an interface item
           contains the words found in this list then key = item in the list
           and value = remaining words of the interface item. If an interface
           item is found then the other methods are not considered.
        2. Portkeys dict is examined. If interface item contains an item
           found in a list of the dict then corresponding key (i.e. 1 or 2)
           is used to split the item. The key of the item is equal to the
           number of words of the dict key, the rest of the item = value.
           Example: If line = channel-group 2 mode active, then
           key = "channel-group"  and value = "2 mode active". If an interface
           item is found then the last method is not considered.
        3. Default method. Last word of line = value
           and all preceding words = key.
        """

        portkey_exceptions = ['ip vrf forwarding']

        portkeys = {1: ['hold-queue', 'standby', 'channel-group', 'description'],
                    2: ['switchport port-security', 'ip', 'spanning-tree',
                        'speed auto', 'srr-queue bandwidth']}

        line = line.lstrip()
        found_item = False

        # 'Method 1'
        for key in portkey_exceptions:
            if key in line:
                if 'Vlan' in portindex:
                    switch['vlan'][vlanindex][key] = get_value(key, line)
                    found_item = True
                else:
                    switch['port'][portindex][key] = get_value(key, line)
                    found_item = True

        # 'Method 2'
        for key_length in portkeys:
            if found_item:
                continue
            for item in portkeys[key_length]:
                if item not in line:
                    continue
                key = get_key(line, key_length)
                if 'standby' in line:
                    if 'Vlan' in portindex:
                        standby.append(get_value(key, line))
                        switch['vlan'][vlanindex]['standby'] = ','.join(standby)
                        found_item = True
                    else:
                        standby.append(get_value(key, line))
                        switch['port'][portindex]['standby'] = ','.join(standby)
                        found_item = True
                elif 'ip helper-address' in line:
                    if 'Vlan' in portindex:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        switch['vlan'][vlanindex]['ip helper-address'] = helper
                        found_item = True
                    else:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        switch['port'][portindex]['ip helper-address'] = helper
                        found_item = True
                elif 'Vlan' in portindex:
                    switch['vlan'][vlanindex][key] = get_value(key, line)
                    found_item = True
                else:
                    switch['port'][portindex][key] = get_value(key, line)
                    found_item = True

        # 'Method 3 or default method'
        if not found_item:
            key = get_key(line, len(line.split())-1)
            if 'switchport trunk allowed vlan' in line:
                for raw_vlans in get_value(key, line).split(','):
                    if '-' in raw_vlans:
                        for vlan_id in splitrange(raw_vlans):
                            vlan_allow_list.append(vlan_id)
                    else:
                        vlan_allow_list.append(raw_vlans)
                allow_lst = ','.join(vlan_allow_list)
                switch['port'][portindex]['vlan_allow_list'] = allow_lst
            elif 'Vlan' in portindex:
                switch['vlan'][vlanindex][key] = get_value(key, line)
            else:
                switch['port'][portindex][key] = get_value(key, line)


    # Start main part of function
    switchinfo = [] # list of switch objects
    
    for configfile in configfiles:

        switch = Vividict()

        with open(configfile, 'r') as f:
            lines = f.readlines()

        context = ''
        match = ReSearcher()

        for line in lines:

            line = line.rstrip()

            if match(r'^hostname (.*)', line):
                hostname = format(match.group(1))
                switch['hostname'] = hostname
 
            elif match(r'^interface (Vlan(\d+))', line):
                context = 'port'
                portindex = format(match.group(1))
                vlanindex = format(match.group(2))
                switch['vlan'][vlanindex]['vlanindex'] = vlanindex
                standby = []
                ip_helper = []

            elif match(r'^interface (.*)', line):
                context = 'port'
                portindex = format(match.group(1))
                switch['port'][portindex]['portindex'] = portindex
                vlan_allow_list = []
                standby = []
                ip_helper = []

            elif match(r'^vlan (\d+)$', line):
                context = 'vlan'
                vlanindex = format(match.group(1))
                switch['vlan'][vlanindex]['vlanindex'] = vlanindex

            elif match(r'^vlan ([0-9,-]+)', line):
                context = 'vlan'
                value = format(match.group(1))
                for raw_vlans in value.split(','):
                    if '-' in raw_vlans:
                        for vlan in splitrange(raw_vlans):
                            switch['vlan'][str(vlan)]['vlanindex'] = str(vlan)
                    else:
                        switch['vlan'][raw_vlans]['vlanindex'] = str(raw_vlans)
                       
            elif context == 'port':

                if match(r'^ no (.*)', line):
                    key = format(match.group(1))
                    value = format(match.group(0))
                    if 'Vlan' in portindex:
                        switch['vlan'][vlanindex][key] = value
                    else:
                        switch['port'][portindex][key] = value

                # interface items are stored with helper function
                elif match('^ .*', line):
                    store_port_items(line, vlanindex, portindex)

                elif match(r'!$', line):
                    context = ''

            elif context == 'vlan':

                if match(r'^ name (.*)', line):
                    switch['vlan'][vlanindex]['name'] = format(match.group(1))

                elif match(r'!$', line):
                    context = ''
       
        switchinfo.append(switch)

    return switchinfo



def info_to_xls(switchinfo):

    """
    Function print Switchinfo dictionaty to excel file. Vlan
    are port info are printed in seperated tabs.
    """

    # Calculate set of keys.
    vlankeys = set()
    portkeys = set()

    for switch in switchinfo:
        for vlan, vlanitems in switch['vlan'].items():
            vlankeys.update(vlanitems.keys())
               
    for switch in switchinfo:
        for port, portitems in switch['port'].items():
            portkeys.update(portitems.keys())
                
    vlankeys = sorted(set(vlankeys))
    portkeys = sorted(set(portkeys))
    if 'vlanindex' in vlankeys:
        vlankeys.remove('vlanindex')
    if 'name' in vlankeys:
        vlankeys.remove('name')
    if 'description'in portkeys:
        portkeys.remove('description')
    if 'portindex'in portkeys:
        portkeys.remove('portindex')
    vlankeys.insert(0, 'name')
    portkeys.insert(0, 'description')

    wb = Workbook()
    ws = wb.create_sheet("Vlaninfo", 0)
    ws = wb.create_sheet("Portinfo", 0)

    ws = wb['Vlaninfo']

    count_vlan_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'vlanindex'
    for count, vlankey in enumerate(vlankeys):
        ws[xlref(0, count+2)] = vlankey

    for switch in switchinfo:
        for vlan, vlanitems in switch['vlan'].items():
            ws[xlref(count_vlan_row+1, 0)] = switch['hostname']
            ws[xlref(count_vlan_row+1, 1)] = vlan

            for count_col, vlankey in enumerate(vlankeys):
                value = vlanitems.get(vlankey, '')
                ws[xlref(count_vlan_row+1, count_col+2)] = value
            count_vlan_row += 1

    ws = wb['Portinfo']

    count_port_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'interface'
    for count, portkey in enumerate(portkeys):
        ws[xlref(0, count+2)] = portkey

    for switch in switchinfo:
        for port, portitems in switch['port'].items():
            ws[xlref(count_port_row+1, 0)] = switch['hostname']
            ws[xlref(count_port_row+1, 1)] = port

            for count_col, portkey in enumerate(portkeys):
                value = portitems.get(portkey, '')
                ws[xlref(count_port_row+1, count_col+2)] = value
            count_port_row += 1

    wb.save('result.xlsx')


if __name__ == "__main__":

    #os.chdir('C:/Users/Hans/Desktop/GIT/SwitchInterfaceParser')

    configfiles = [file for file in glob.glob('*.txt')]

    # Retrieve interface and vlan info from configuration files
    # and store in switchinfo dictionary.
    switchinfo = get_switch_info(configfiles)

    # Print Switchinfo object in excel file.
    info_to_xls(switchinfo)
